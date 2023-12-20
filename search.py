from tkinter import *
import tkinter as tk
import openpyxl, xlrd
from openpyxl import Workbook

main = Tk()
main.title("Procurar")
main.geometry("800x500")
main.config(highlightbackground="black", highlightthickness=2)

excel_path = 'E:\\VS Code\\Python_Projetos\\Search\\Notas.xlsx'

def submit():
    search_empid = numero.get()
    search_name = nome.get()
    
    numero.configure(state=tk.NORMAL)
    nome.configure(state=tk.NORMAL)
    senha.configure(state=tk.NORMAL)
    nota.configure(state=tk.NORMAL)
    
    numero.delete(0,'end')
    nome.delete(0,'end')
    senha.delete(0,'end')
    nota.delete(0,'end')
    
    file = openpyxl.load_workbook(excel_path)
    sheet = file['Planilha1']
    
    for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True):
        if cell[0] == int(search_empid):
        #if cell[1] == str(search_name):
            numero.insert(0, cell[0])
            nome.insert(0, cell[1])
            senha.insert(0, cell[2])
            nota.insert(0, cell[3])
            
            numero.configure(state=tk.DISABLED)
            nome.configure(state=tk.DISABLED)
            senha.configure(state=tk.DISABLED)
            nota.configure(state=tk.DISABLED)

frame1 = LabelFrame(main, text="Search:").pack(expand='yes', fill='both')

Label(frame1, text="Número: ").place(x=50,y=60)
Label(frame1, text="Nome: ").place(x=50,y=100)
Label(frame1, text="Senha: ").place(x=50,y=130)
Label(frame1, text="Nota: ").place(x=50,y=160)

numero = Entry(frame1)
numero.place(x=250,y=60)
nome = Entry(frame1)
nome.place(x=250,y=100)
senha = Entry(frame1)
senha.place(x=250,y=130)
nota = Entry(frame1)
nota.place(x=250,y=160)

#nome.configure(state=tk.DISABLED)
senha.configure(state=tk.DISABLED)
nota.configure(state=tk.DISABLED)

Button(frame1, text="Procurar", command=submit).place(x=400,y=200)

main.mainloop()


            
    
    
    
    
    
    


    
